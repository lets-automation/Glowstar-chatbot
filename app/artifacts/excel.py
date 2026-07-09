"""
excel.py
--------
Turn query results into an Excel (.xlsx) file using openpyxl.

Input shape (same as the query runner returns):
  columns: ["Shape", "Cnt", ...]
  rows:    [ {"Shape": "ROUND", "Cnt": 123}, ... ]
"""

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font
from openpyxl.worksheet.properties import PageSetupProperties

from app.artifacts import output_path
from app.artifacts.charts import CHART_MAX_ROWS, pick_chart_columns

# One clean brand colour for every bar (no rainbow). Matches the app's accent.
_BAR_COLOR = "2A78D6"


# Leading characters that spreadsheet apps interpret as the start of a FORMULA.
# A DB/user value like "=1+1" or "=cmd|'/c calc'!A0" would otherwise execute on
# open (CSV/DDE formula injection), so we neutralise them with a leading quote.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _clean(value):
    """
    Make a value safe to write to a cell:
    - strip XML-illegal control characters (legacy ERP free-text fields contain
      control bytes 0x00-0x1F that openpyxl rejects with IllegalCharacterError,
      crashing the whole export);
    - defuse FORMULA INJECTION: a string beginning with = + - @ (or tab/CR) is
      prefixed with a single quote so the spreadsheet treats it as literal text,
      never an executable formula.
    """
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if value[:1] in _FORMULA_TRIGGERS:
            value = "'" + value
    return value


def _add_chart_sheet(wb, data_ws, columns: list[str], rows: list[dict]) -> None:
    """
    Add a second 'Chart' sheet with a native Excel bar chart built from the
    data sheet - only when the data is chartable and not too large. The chart
    is a real editable Excel chart referencing the data cells.
    """
    if not (1 < len(rows) <= CHART_MAX_ROWS):
        return
    picked = pick_chart_columns(columns, rows)
    if not picked:
        return
    label_col, value_col = picked
    label_idx = columns.index(label_col) + 1  # 1-based column numbers
    value_idx = columns.index(value_col) + 1
    last_row = len(rows) + 1  # +1 for the header row

    chart = BarChart()
    chart.type = "col"
    chart.title = f"{value_col} by {label_col}"
    chart.legend = None
    chart.height = 10
    chart.width = 20             # fits a printed page; all bars stay visible
    chart.gapWidth = 55          # fuller bars, tighter gaps
    chart.varyColors = False     # all bars one colour, not a rainbow

    # Include the header cell in the data range so the series is named.
    data_ref = Reference(data_ws, min_col=value_idx, min_row=1, max_row=last_row)
    cats_ref = Reference(data_ws, min_col=label_idx, min_row=2, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # CRITICAL: openpyxl marks both axes 'deleted' by default, which makes Excel
    # hide the category labels (the D/E/F/G under each bar) and the value scale.
    # Setting delete=False makes them show - this is the main fix.
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    # NO axis TITLES: openpyxl can't reliably position them and Excel renders
    # them on top of the tick labels ("FluorescentStones" over the numbers,
    # "Color" over D/G/H). The chart title + the tick labels already say it all.
    chart.x_axis.title = None
    chart.y_axis.title = None
    chart.x_axis.tickLblPos = "low"      # category labels along the bottom
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.numFmt = "#,##0"        # thousands separators on the value axis
    # Soft light-grey gridlines instead of the heavy default black.
    chart.y_axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(solidFill="D9D9D9", w=9525))
    )
    chart.x_axis.majorGridlines = None

    # One clean colour for all bars.
    series = chart.series[0]
    series.graphicalProperties.solidFill = _BAR_COLOR
    series.graphicalProperties.line.solidFill = _BAR_COLOR

    # Data labels: show ONLY the value, sitting just above each bar. The series
    # name and category name MUST be off - leaving them on prints
    # "FluorescentStones, D, 8918" on every bar, which overlaps into a mess.
    dlbls = DataLabelList()
    dlbls.showVal = True
    dlbls.showSerName = False
    dlbls.showCatName = False
    dlbls.showLegendKey = False
    dlbls.showPercent = False
    dlbls.showBubbleSize = False
    dlbls.numFmt = "#,##0"
    dlbls.dLblPos = "outEnd"   # above the top of each column
    series.dLbls = dlbls

    chart_ws = wb.create_sheet("Chart")
    chart_ws.add_chart(chart, "B2")
    # Print the chart sheet in landscape, scaled to fit one page - so if anyone
    # prints or PDFs it, the whole chart shows instead of being clipped.
    chart_ws.page_setup.orientation = "landscape"
    chart_ws.page_setup.fitToWidth = 1
    chart_ws.page_setup.fitToHeight = 1
    chart_ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def to_excel(columns: list[str], rows: list[dict], filename: str = "report.xlsx") -> str:
    """
    Write columns + rows to an .xlsx file. The numbers go on a 'Data' sheet; if
    the data is chartable, a second 'Chart' sheet holds a bar chart of it.
    Returns the file path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header row (bold).
    ws.append([_clean(c) for c in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows, in column order (control chars stripped so openpyxl can't choke).
    for row in rows:
        ws.append([_clean(row.get(col)) for col in columns])

    # Auto-ish column widths for readability.
    for i, col in enumerate(columns, start=1):
        max_len = max([len(str(col))] + [len(str(r.get(col, ""))) for r in rows])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(
            max_len + 2, 50
        )

    # Second sheet: a chart of the data (when chartable).
    _add_chart_sheet(wb, ws, columns, rows)

    path = output_path(filename)
    wb.save(path)
    return path
