"""
test_multi_section_export.py
----------------------------
A "full report" question runs several queries - production, damage, bonus, GIA -
and the chat answer narrates all of them. The Excel download used to carry only
ONE, because the export was built from the single result_capture winner.

Client report: asked for a full report, saw bonus + GIA + production + damage in
the chat, downloaded a workbook with one sheet of 318 production rows and
nothing else. Verified against their file: SHEETS: ['Data'].
"""
import openpyxl
import pytest

from app.agent.result_capture import add_section
from app.artifacts.excel import to_excel_sections

PRODUCTION = (["KapanName", "PacketNo", "Carats"],
              [{"KapanName": "NS26", "PacketNo": i, "Carats": 1.1} for i in range(5)])
BONUS = (["EmployeeName", "BonusPoints"],
         [{"EmployeeName": "PANELIYA SANJAY", "BonusPoints": 120}])
DAMAGE = (["KapanName", "DamagePoints"], [{"KapanName": "NS26", "DamagePoints": 3}])
LOOKUP = (["DepartMentName"], [{"DepartMentName": "MFG - 1"}])


def _sections(*pairs):
    out: list[dict] = []
    for cols, rows in pairs:
        add_section(out, cols, rows)
    return out


def test_every_section_is_kept():
    assert len(_sections(PRODUCTION, BONUS, DAMAGE)) == 3


def test_lookups_are_not_exported_as_sections():
    # A one-column "how is this spelled" query is not a report section.
    assert len(_sections(PRODUCTION, LOOKUP)) == 1


def test_empty_results_are_skipped():
    assert _sections((["A", "B"], [])) == []


def test_a_rerun_of_the_same_query_is_not_duplicated():
    # Models re-run a query after a corrective nudge; two identical sheets read
    # as a mistake to whoever opens the file.
    assert len(_sections(PRODUCTION, PRODUCTION)) == 1


def test_workbook_has_one_sheet_per_section(tmp_path):
    path = to_excel_sections(_sections(PRODUCTION, BONUS, DAMAGE),
                             filename="test-multi.xlsx")
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) == 3, f"sections were lost: {wb.sheetnames}"

    # Sheet names come from the columns - the model decides what a "full report"
    # contains, so there is no fixed list of section names to map to.
    joined = " ".join(wb.sheetnames)
    assert "BonusPoints" in joined or "EmployeeName" in joined

    # The bonus figure must actually be IN the file, not just named by a tab.
    values = {c.value for ws in wb.worksheets for row in ws.iter_rows() for c in row}
    assert 120 in values, "the bonus section's data is missing from the workbook"


def test_a_single_section_keeps_the_normal_single_result_file():
    """One section must behave exactly as before - a 'Data' sheet, plus the
    existing auto 'Chart' sheet when the data is chartable. Only genuinely
    multi-part reports get the per-section layout."""
    path = to_excel_sections(_sections(PRODUCTION), filename="test-single.xlsx")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames[0] == "Data"
    assert set(wb.sheetnames) <= {"Data", "Chart"}


def test_nothing_to_export_raises_rather_than_writing_an_empty_file():
    with pytest.raises(ValueError):
        to_excel_sections([], filename="test-empty.xlsx")
